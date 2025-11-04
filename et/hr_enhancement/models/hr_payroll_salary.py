from odoo import models, fields, api
from odoo.exceptions import ValidationError
import base64, xlsxwriter, io
from datetime import date, datetime
import base64, io
from openpyxl import load_workbook
from datetime import datetime


class HrPayrollSalary(models.Model):
    _name = 'hr.payroll.salary'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _description = 'Planilla de Sueldo de Empleados'

    name = fields.Char(string="Nombre de la Planilla", required=True, tracking=True)
    date_start = fields.Date(string="Fecha Inicio", required=True, default=fields.Date.today, tracking=True)
    date_end = fields.Date(string="Fecha Fin", required=True, default=fields.Date.today, tracking=True)
    #colocar el tipo de pago
    type_paid = fields.Selection([
        ('semanal', 'Semanal'),
        ('quincenal', 'Quincenal'),
        ('mensual', 'Mensual'),
        ('dia', 'Día'),
        ('otro', 'Otro')
    ], string="Tipo de Pago", default='quincenal', required=True, tracking=True)
    pay_month = fields.Selection([
        ('1', 'Enero'), ('2', 'Febrero'), ('3', 'Marzo'), ('4', 'Abril'), ('5', 'Mayo'), ('6', 'Junio'),
        ('7', 'Julio'), ('8', 'Agosto'), ('9', 'Septiembre'), ('10', 'Octubre'), ('11', 'Noviembre'), ('12', 'Diciembre'),
    ], string='Mes', default=lambda self: str(date.today().month), tracking=True)
    pay_year = fields.Integer(string='Año', tracking=True, default=lambda self: date.today().year)
    #que mas deberia polocar en un payroll
    state = fields.Selection([('draft', 'Borrador'),
                                ('confirmed', 'Confirmado'),
                                ('paid', 'Pagado'),
                                ('partial_paid', 'Parcialmente Pagado'),
                                ('cancelled', 'Cancelado')],
                            string='Estado', default='draft', tracking=True)
    currency_id = fields.Many2one(
        'res.currency',
        string="Moneda",
        required=True,
        default=lambda self: self.env.company.currency_id.id
    )
    employee_type = fields.Selection([
        ('employee', 'Fijo'),
        ('eventual', 'Eventual'),
    ], string="Tipo de Empleado", default='eventual', required=True, tracking=True)
    #totales de la planilla
    total_amount = fields.Monetary(
        string="Total a Pagar",
        compute='_compute_total_amount',
        currency_field='currency_id',
        store=True
    )
    #seleccionar que tipo de eventual o empleado o turno
    type_liquidacion_eventual = fields.Selection([
        ('eventual_day', 'Eventuales de Turno Día'),
        ('eventual_night', 'Eventuales de Turno Noche')
    ], string="Tipo de Liquidación", default='eventual_day', required=True, tracking=True)
    type_liquidacion_employee = fields.Selection([
        ('employee_day', 'Empleado Día'),
        ('employee_night', 'Empleado Noche')
    ], string="Tipo de Liquidación", default='employee_day', required=True, tracking=True)
    line_ids = fields.One2many('hr.payroll.salary.line', 'payroll_id', string="Detalles de Planilla", copy=True)
    import_file = fields.Binary('Archivo Liquidación (XLSX)')
    import_filename = fields.Char('Nombre archivo')
    labor_cost_id = fields.Many2one('hr.season.labor.cost', string="Costo Laboral", help="Costo laboral aplicado a la planilla", store=True)   # como forzar guardado de un dato si esta readonly


    @api.model
    def create(self, vals):
        # Si viene default desde el action, respetalo
        if 'employee_type' not in vals and self.env.context.get('default_employee_type'):
            vals['employee_type'] = self.env.context['default_employee_type']
        rec = super().create(vals)
        return rec

    def write(self, vals):
        # No permitir cambiar employee_type una vez creado
        if 'employee_type' in vals:
            for r in self:
                if r.employee_type and vals['employee_type'] != r.employee_type:
                    raise ValidationError("El tipo de empleado no puede modificarse en la planilla.")
        return super().write(vals)

    @api.depends('line_ids.net_amount')
    def _compute_total_amount(self):
        for rec in self:
            rec.total_amount = sum(rec.line_ids.mapped('net_amount'))
            
    def action_confirm(self):
        for record in self:
            if record.state != 'draft':
                raise ValidationError('Solo se puede confirmar una planilla en estado Borrador.')
            record.state = 'confirmed'
            # Aquí podrías agregar lógica adicional para procesar la planilla, como enviar notificaciones o generar informes.
            
    def action_cancelled(self):
        for record in self:
            if record.state not in ['draft', 'confirmed']:
                raise ValidationError('Solo se puede cancelar una planilla en estado Borrador o Confirmado.')
            record.state = 'cancelled'
            # Aquí podrías agregar lógica adicional para manejar la cancelación, como revertir pagos o notificar a los empleados.

    def action_paid(self):
        for record in self:
            if record.state != 'confirmed':
                raise ValidationError('Solo se puede marcar una planilla como Pagada en estado Confirmado.')
            total = len(record.line_ids)
            pagadas = len(record.line_ids.filtered('is_paid'))
            if pagadas == 0:
                record.state = 'confirmed'
            elif pagadas < total:
                record.state = 'partial_paid'
            else:
                record.state = 'paid'
            # Aquí podrías agregar lógica adicional para procesar el pago, como registrar transacciones contables o enviar notificaciones a los empleados.

    def action_draft(self):
        for record in self:
            if record.state not in ['confirmed']:
                raise ValidationError('Solo se puede restablecer a Borrador una planilla en estado Confirmado.')
            record.state = 'draft'
            
    def action_load_employees(self):
        for record in self:
            if record.state != 'draft':
                raise ValidationError('Solo se pueden cargar empleados en una planilla en estado Borrador.')
            # no eliminar las lineas de la planilla sino traer los empleados que falten
            if record.line_ids:
                existing_employees = record.line_ids.mapped('employee_id')
                domain = [('employee_type', '=', record.employee_type), ('id', 'not in', existing_employees.ids)]
                if record.employee_type == 'eventual':
                    if record.type_liquidacion_eventual == 'eventual_day':
                        domain.append(('type_shift', '=', 'day'))
                    else:
                        domain.append(('type_shift', '=', 'night'))
                elif record.employee_type == 'employee':
                    if record.type_liquidacion_employee == 'employee_day':
                        domain.append(('type_shift', '=', 'day'))
                    else:
                        domain.append(('type_shift', '=', 'night'))
                employees = self.env['hr.employee'].search(domain)
            else:
                domain = [('employee_type', '=', record.employee_type)]
                if record.employee_type == 'eventual':
                    if record.type_liquidacion_eventual == 'eventual_day':
                        domain.append(('type_shift', '=', 'day'))
                    else:
                        domain.append(('type_shift', '=', 'night'))
                elif record.employee_type == 'employee':
                    if record.type_liquidacion_employee == 'employee_day':
                        domain.append(('type_shift', '=', 'day'))
                    else:
                        domain.append(('type_shift', '=', 'night'))
                employees = self.env['hr.employee'].search(domain)
            for emp in employees:
                attendances = self.env['hr.attendance'].search([
                    ('employee_id', '=', emp.id),
                    ('check_in', '>=', record.date_start),
                    ('check_in', '<=', record.date_end),
                ])
                if attendances:
                    total_hours = sum(a.worked_hours for a in attendances)
                    total_overtime = sum(a.overtime for a in attendances)
                    total_holiday_hours = sum(a.holiday_hours for a in attendances)
                    self.env['hr.payroll.salary.line'].create({
                        'payroll_id': record.id,
                        'employee_id': emp.id,
                        'basic_amount': emp.wage,
                        'worked_hours': total_hours,
                        'overtime': total_overtime,
                        'holiday_hours': total_holiday_hours,
                    })
                    
    def action_load_employees(self):
        for record in self:
            if record.state != 'draft':
                raise ValidationError('Solo se pueden cargar empleados en una planilla en estado Borrador.')

            # 1) buscar planillas que se solapen en fechas y sean del mismo tipo/turno
            overlapped = self.env['hr.payroll.salary'].search([
                ('id', '!=', record.id),
                ('employee_type', '=', record.employee_type),
                # mismo subtipo de liquidación
                '|',
                    '&', ('employee_type', '=', 'eventual'),
                        ('type_liquidacion_eventual', '=', record.type_liquidacion_eventual),
                    '&', ('employee_type', '=', 'employee'),
                        ('type_liquidacion_employee', '=', record.type_liquidacion_employee),
                # solapamiento de fechas
                ('date_start', '<=', record.date_end),
                ('date_end', '>=', record.date_start),
                # estados que indican que ya se liquidó (total o parcial)
                ('state', 'in', ['partial_paid', 'paid'])
            ])

            # 2) empleados ya liquidados en ese solape
            already_paid_emps = self.env['hr.payroll.salary.line'].search([
                ('payroll_id', 'in', overlapped.ids),
                ('is_paid', '=', True),
            ]).mapped('employee_id')

            # 3) base domain por tipo y turno de la planilla actual
            domain = [('employee_type', '=', record.employee_type),
                    ('id', 'not in', already_paid_emps.ids)]
            if record.line_ids:
                domain.append(('id', 'not in', record.line_ids.mapped('employee_id').ids))

            # turno
            if record.employee_type == 'eventual':
                domain.append(('type_shift', '=', 'day' if record.type_liquidacion_eventual == 'eventual_day' else 'night'))
            else:
                domain.append(('type_shift', '=', 'day' if record.type_liquidacion_employee == 'employee_day' else 'night'))

            employees = self.env['hr.employee'].search(domain)

            # 4) crear líneas solo para los no-liquidados
            for emp in employees:
                attendances = self.env['hr.attendance'].search([
                    ('employee_id', '=', emp.id),
                    ('check_in', '>=', record.date_start),
                    ('check_in', '<=', record.date_end),
                    ('check_out', '!=', False),
                ])
                if attendances:
                    total_hours = sum(a.worked_hours for a in attendances)
                    total_overtime = sum(a.overtime for a in attendances)
                    total_holiday_hours = sum(a.holiday_hours for a in attendances)
                    self.env['hr.payroll.salary.line'].create({
                        'payroll_id': record.id,
                        'employee_id': emp.id,
                        'basic_amount': emp.wage,
                        'worked_hours': total_hours,
                        'overtime': total_overtime,
                        'holiday_hours': total_holiday_hours,
                    })


    def action_clear_lines(self):
        for record in self:
            if record.state != 'draft':
                raise ValidationError('Solo se pueden eliminar las líneas de una planilla en estado Borrador.')
            record.line_ids.unlink()
            
    @api.onchange('type_liquidacion_eventual')
    def _onchange_type_liquidacion_eventual_clear_line(self):
        for record in self:
            record.action_clear_lines()

    @api.onchange('date_start', 'date_end')
    def _onchange_cost_laboral(self):
        for rec in self:
            season_costo = self.env['hr.season.labor.cost'].search([('state', '=', 'active'),
                                                                    ('date_start', '<=', rec.date_start),
                                                                    ('date_end', '>=', rec.date_end)], limit=1)
            if not season_costo:
                raise ValidationError('No hay una temporada activa para calcular el costo laboral. Por favor, cree y active una temporada en Costo Laboral por Temporada.')
            rec.labor_cost_id = season_costo.id

    def action_generar_excel(self):
        self.ensure_one()
        # Validaciones básicas
        if not self.line_ids:
            raise ValidationError("La planilla no tiene líneas para exportar.")

        # Buffer memoria
        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output, {'in_memory': True})
        ws = wb.add_worksheet('Planilla')

        # ==== formatos ====
        fmt_header = wb.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'bg_color': '#D9E1F2', 'border': 1
        })
        fmt_left = wb.add_format({'align': 'left', 'valign': 'vcenter', 'border': 1})
        fmt_right = wb.add_format({'align': 'right', 'valign': 'vcenter', 'border': 1})
        fmt_float2 = wb.add_format({'num_format': '0', 'align': 'right', 'valign': 'vcenter', 'border': 1})
        fmt_float3 = wb.add_format({'num_format': '0', 'align': 'right', 'valign': 'vcenter', 'border': 1, 'bold': True})
        fmt_money = wb.add_format({'num_format': '#,##0.00', 'align': 'right', 'valign': 'vcenter', 'border': 1,})
        fmt_money2 = wb.add_format({'num_format': '#,##0.00', 'align': 'right', 'valign': 'vcenter', 'border': 1, 'bold': True})
        fmt_title = wb.add_format({'bold': True, 'font_size': 14})

        # ==== título / cabecera de planilla ====
        ws.write(0, 0, 'Planilla', fmt_title)
        ws.write(1, 0, 'Nombre')
        ws.write(1, 1, self.name or '')
        ws.write(2, 0, 'Tipo Pago')
        ws.write(2, 1, dict(self._fields['type_paid'].selection).get(self.type_paid, ''))
        ws.write(3, 0, 'Tipo Empleado')
        ws.write(3, 1, dict(self._fields['employee_type'].selection).get(self.employee_type, ''))
        ws.write(4, 0, 'Período')
        periodo = ""
        if self.type_paid == 'mensual':
            periodo = f"{self.pay_month}/{self.pay_year}"
        else:
            periodo = f"{self.date_start or ''} al {self.date_end or ''}"
        ws.write(4, 1, periodo)
        ws.write(5, 0, 'Estado')
        ws.write(5, 1, dict(self._fields['state'].selection).get(self.state, ''))

        start_row = 7

        # ==== anchos de columna ====
        ws.set_column(0, 0, 15)    # #
        ws.set_column(1, 1, 40)   # Empleado
        ws.set_column(2, 2, 10)   # Dias trabajadas
        ws.set_column(3, 3, 10)   # Horas trabajadas
        ws.set_column(4, 4, 10)   # 50%
        ws.set_column(5, 5, 10)   # 100%
        ws.set_column(6, 6, 15)   # Monto H. Trabajadas
        ws.set_column(7, 7, 15)   # Monto 50%
        ws.set_column(8, 8, 15)   # Monto 100%
        ws.set_column(9, 9, 15)   # Monto Neto
        ws.set_column(10, 10, 15)   # Bono
        ws.set_column(11, 11, 15) # Descuento
        ws.set_column(12, 12, 15) # Total a cobrar
        ws.set_column(13, 13, 10) # Pagado
        ws.set_column(14, 14, 35) # Nota

        # ==== encabezados ====
        headers = [
            'N°', 'EMPLEADO', 'DÍAS TRAB.', 'H. TRAB.', 'H. 50%', 'H. 100%', 'MONTO H. TRAB.', 'MONTO 50%', 'MONTO 100%', 'MONTO NETO', 'BONO', 'DESCUENTO', 'TOTAL A COBRAR', '¿PAGADO?', 'NOTA'
        ]

        for c, h in enumerate(headers):
            ws.write(start_row, c, h, fmt_header)

        row = start_row + 1
        i = 1

        # Totales
        tot_worked = tot_overtime = tot_holiday = tot_days_worked = 0.0
        tot_worked_hours_amount = tot_overtime_amount = tot_holiday_hours_amount = 0.0
        tot_basic = tot_bonus = tot_discount = tot_gross = tot_net = 0.0

        for line in self.line_ids:
            ws.write_number(row, 0, i, fmt_right)
            ws.write(row, 1, line.employee_id.name or '', fmt_left)
            ws.write_number(row, 2, line.worked_days or 0.0, fmt_float2)
            ws.write_number(row, 3, line.worked_hours or 0.0, fmt_float2)
            ws.write_number(row, 4, line.overtime or 0.0, fmt_float2)
            ws.write_number(row, 5, line.holiday_hours or 0.0, fmt_float2)
            ws.write_number(row, 6, line.worked_hours_amount or 0.0, fmt_money)
            ws.write_number(row, 7, line.overtime_amount or 0.0, fmt_money)
            ws.write_number(row, 8, line.holiday_hours_amount or 0.0, fmt_money)
            ws.write_number(row, 9, line.gross_amount or 0.0, fmt_money)
            ws.write_number(row, 10, line.bonus or 0.0, fmt_money)
            ws.write_number(row, 11, line.discount or 0.0, fmt_money)
            ws.write_number(row, 12, line.net_amount or 0.0, fmt_money)
            ws.write(row, 13, 'Sí' if line.is_paid else 'No', fmt_left)
            ws.write(row, 14, line.note or '', fmt_left)

            # acumula totales
            tot_days_worked += (line.worked_days or 0.0)
            tot_worked += (line.worked_hours or 0.0)
            tot_overtime += (line.overtime or 0.0)
            tot_holiday += (line.holiday_hours or 0.0)

            tot_worked_hours_amount += (line.worked_hours_amount or 0.0)
            tot_overtime_amount += (line.overtime_amount or 0.0)
            tot_holiday_hours_amount += (line.holiday_hours_amount or 0.0)
            tot_bonus += (line.bonus or 0.0)
            tot_discount += (line.discount or 0.0)
            tot_gross += (line.gross_amount or 0.0)
            tot_net += (line.net_amount or 0.0)

            row += 1
            i += 1

        # ==== fila de totales ====
        ws.merge_range(row, 0, row, 1, "TOTALES", fmt_header)
        ws.write_number(row, 2, tot_days_worked, fmt_float3)
        ws.write_number(row, 3, tot_worked, fmt_float3)
        ws.write_number(row, 4, tot_overtime, fmt_float3)
        ws.write_number(row, 5, tot_holiday, fmt_float3)
        ws.write_number(row, 6, tot_worked_hours_amount, fmt_money2)
        ws.write_number(row, 7, tot_overtime_amount, fmt_money2)
        ws.write_number(row, 8, tot_holiday_hours_amount, fmt_money2)
        ws.write_number(row, 9, tot_gross, fmt_money2)
        ws.write_number(row, 10, tot_bonus, fmt_money2)
        ws.write_number(row, 11, tot_discount, fmt_money2)
        ws.write_number(row, 12, tot_net, fmt_money2)
        ws.merge_range(row, 13, row, 14, " ", fmt_header)
        wb.close()
        output.seek(0)

        # adjunto
        filename = f'Planilla_{self.name or "sin_nombre"}_{datetime.today().strftime("%Y%m%d")}.xlsx'
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })

        # descarga
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }


    def action_importar_liquidacion(self):
        self.ensure_one()
        if self.state not in ('confirmed', 'paid'):
            raise ValidationError("La planilla debe estar Confirmada para importar liquidaciones.")
        if not self.import_file:
            raise ValidationError("Subí un archivo XLSX en el campo 'Archivo Liquidación'.")

        # Leer XLSX
        data = base64.b64decode(self.import_file)
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active  # o por nombre si lo fijaste

        # Ubicar índices de columnas por encabezado
        header_row = None
        cols = {}
        needed = {'LINE_ID', 'LIQUIDADO'}
        for r in range(1, 30):  # busca encabezado en primeras 30 filas
            vals = [ws.cell(row=r, column=c).value for c in range(1, 50)]
            if vals and any(vals):
                # normalizar
                for idx, v in enumerate(vals, start=1):
                    nv = (str(v).strip().upper() if v is not None else '')
                    if nv in {'LINE_ID','LIQUIDADO'}:
                        cols[nv] = idx
                if needed.issubset(set(cols)):
                    header_row = r
                    break
        if not header_row:
            raise ValidationError("No se encontraron columnas 'LINE_ID' y 'LIQUIDADO' en el archivo.")

        # Marcar líneas pagadas según XLSX
        Line = self.env['hr.payroll.salary.line'].sudo()
        marcados = 0
        for r in range(header_row+1, ws.max_row+1):
            line_id = ws.cell(row=r, column=cols['LINE_ID']).value
            liq_val = ws.cell(row=r, column=cols['LIQUIDADO']).value
            if not line_id:
                continue
            flag = str(liq_val).strip().lower() if liq_val is not None else ''
            if flag in {'l','x','1','si','sí','yes','y'}:
                line = Line.search([('id','=', int(line_id)), ('payroll_id','=', self.id)], limit=1)
                if line and not line.is_paid:
                    line.write({
                        'is_paid': True,
                        'user_paid': self.env.user.id,
                        'date_paid': fields.Date.today(),
                    })
                    marcados += 1

        # Actualizar estado de la planilla
        total = len(self.line_ids)
        pagadas = len(self.line_ids.filtered(lambda l: l.is_paid))
        if pagadas == 0:
            # permanece en confirmed
            pass
        elif 0 < pagadas < total:
            # si querés estado intermedio, agregá 'partial' en tu selección de state
            # self.state = 'partial'
            self.message_post(body=f"Liquidación parcial: {pagadas}/{total} líneas marcadas.")
        else:
            # todas pagadas
            self.state = 'paid'
            self.message_post(body="Todas las líneas fueron marcadas como pagadas.")

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {'title': 'Importación completada',
                       'message': f'Líneas marcadas: {marcados}',
                       'sticky': False}
        }
    
class HrPayrollSalaryLine(models.Model):
    _name = 'hr.payroll.salary.line'
    _description = 'Detalle de Planilla de Pago' 
    _inherit = ['mail.thread', 'mail.activity.mixin']

    payroll_id = fields.Many2one('hr.payroll.salary', string="Planilla", required=True, ondelete='cascade')
    employee_id = fields.Many2one('hr.employee', string="Empleado", required=True)
    employee_state = fields.Selection(
        related='employee_id.state', string='Estado Empleado', store=False, readonly=True
    )
    job_id = fields.Many2one(related='employee_id.job_id', string="Puesto", store=True)
    worked_days = fields.Float('Días Trabajados', compute='_compute_attendance',  readonly=True, store=True)
    worked_hours = fields.Float('H. Trabajadas', compute='_compute_attendance', readonly=True, store=True)
    overtime = fields.Float('H. 50%', compute='_compute_attendance', readonly=True, store=True)
    holiday_hours = fields.Float('H. 100%', compute='_compute_attendance', readonly=True, store=True)
    salary_id = fields.Many2one('hr.employee.salary', string="Sueldo Básico", help="Sueldo básico por el período")
    basic_amount = fields.Monetary('Sueldo Básico', help="Sueldo básico")
    bonus = fields.Monetary('Bonos/Gratificación', currency_field='currency_id', default=0.0)
    discount = fields.Monetary('Descuento', currency_field='currency_id', default=0.0)
    holidays = fields.Monetary('Días de Licencia', currency_field='currency_id', default=0.0)
    worked_hours_amount = fields.Monetary('Monto H. Trabajadas', compute='_compute_amount', currency_field='currency_id', default=0.0, readonly=True)
    overtime_amount = fields.Monetary('Monto al 50%', compute='_compute_amount', currency_field='currency_id', default=0.0, readonly=True)
    holiday_hours_amount = fields.Monetary('Monto al 100%', compute='_compute_amount', currency_field='currency_id', default=0.0, readonly=True)
    gross_amount = fields.Monetary('Monto Neto', currency_field='currency_id', compute='_compute_amount', store=True)
    net_amount = fields.Monetary('Total a Cobrar', currency_field='currency_id', compute='_compute_amount', readonly=True, store=True)
    note = fields.Char('Nota (Observaciones)')
    currency_id = fields.Many2one(
    'res.currency',
    string="Moneda",
    required=True,
    default=lambda self: self.env.company.currency_id.id
    )
    state = fields.Selection(related='payroll_id.state', string='Estado', store=True)
    labor_cost_id = fields.Many2one('hr.season.labor.cost', string="Costo Laboral")
    #campo para saber que si fue pagado para eventuales
    is_paid = fields.Boolean(string="Liquidado", default=False)
    user_paid = fields.Many2one('res.users', string="Usuario que Pagó", readonly=True)
    date_paid = fields.Date(string="Fecha de Pago", readonly=True)
    paid_label = fields.Selection(
        string='Estado Pago',
        selection=[('liquidado', 'Liquidado'), ('sin_liquidar', 'Sin Liquidar')],
        compute='_compute_paid_label'
    )
    @api.depends('is_paid')
    def _compute_paid_label(self):
        for r in self:
            r.paid_label = 'liquidado' if r.is_paid else 'sin_liquidar'

    @api.depends('worked_hours', 'overtime', 'holiday_hours', 'bonus', 'discount')
    def _compute_amount(self):
        for rec in self:
            rec.gross_amount = 0.00
            if rec.labor_cost_id:
                if rec.payroll_id.employee_type == 'eventual' and rec.payroll_id.type_liquidacion_eventual == 'eventual_day':
                    amount_overtime = rec.labor_cost_id.hour_cost_extra
                    amount_holiday = rec.labor_cost_id.hour_cost_holiday
                    rec.gross_amount = ((rec.worked_hours * rec.labor_cost_id.hour_cost_normal) +
                                        (rec.overtime * amount_overtime) +
                                        (rec.holiday_hours * amount_holiday))
                    rec.net_amount = rec.bonus + rec.gross_amount - rec.discount
                    rec.worked_hours_amount = rec.worked_hours * rec.labor_cost_id.hour_cost_normal
                    rec.overtime_amount = rec.overtime * amount_overtime
                    rec.holiday_hours_amount = rec.holiday_hours * amount_holiday
                elif rec.payroll_id.employee_type == 'eventual' and rec.payroll_id.type_liquidacion_eventual == 'eventual_night':
                    amount_overtime = rec.labor_cost_id.hour_cost_extra
                    amount_holiday = rec.labor_cost_id.hour_cost_holiday
                    rec.worked_hours_amount = rec.worked_hours * rec.labor_cost_id.hour_cost_night
                    rec.overtime_amount = rec.overtime * amount_overtime
                    rec.holiday_hours_amount = rec.holiday_hours * amount_holiday
                    rec.gross_amount = rec.worked_hours_amount + rec.overtime_amount + rec.holiday_hours_amount
                    rec.net_amount = rec.bonus + rec.gross_amount - rec.discount
            
    @api.depends('employee_id', 'payroll_id.date_start', 'payroll_id.date_end')
    def _compute_attendance(self):
        for rec in self:
            rec.worked_days = 0.0
            rec.worked_hours = 0.0
            rec.overtime = 0.0
            rec.holiday_hours = 0.0
            season_costo = self.env['hr.season.labor.cost'].search([('state', '=', 'active'),
                                                                    ('date_start', '<=', rec.payroll_id.date_start),
                                                                    ('date_end', '>=', rec.payroll_id.date_end)], limit=1)
            if not season_costo:
                raise ValidationError('No hay una temporada activa para calcular el costo laboral. Por favor, cree y active una temporada en Costo Laboral por Temporada.')
            rec.labor_cost_id = season_costo.id

            if rec.employee_id and rec.payroll_id:
                attendances = self.env['hr.attendance'].search([
                    ('employee_id', '=', rec.employee_id.id),
                    ('check_in', '>=', rec.payroll_id.date_start),
                    ('check_in', '<=', rec.payroll_id.date_end),
                    ('check_out', '!=', False),
                ])
                if  attendances:
                    # Calcular horas trabajadas, días trabajados, horas extras y horas de vacaciones
                    for att in attendances:
                        rec.worked_days += 1
                        rec.worked_hours += att.worked_hours
                        rec.overtime += att.overtime
                        rec.holiday_hours += att.holiday_hours
                else:
                    raise ValidationError(f'No se encontraron registros de asistencia para el empleado {rec.employee_id.name} en el período de la planilla.')

    def action_paid(self):
        self.ensure_one()
        self.write({
            'is_paid': True,
            'user_paid': self.env.user.id,
            'date_paid': fields.Date.today()
        })
        
    def action_cancel_paid(self):
        self.ensure_one()
        self.write({
            'is_paid': False,
            'user_paid': False,
            'date_paid': False
        })
        
    @api.constrains('payroll_id', 'employee_id')
    def _check_unique_employee_in_payroll(self):
        for rec in self:
            dup = self.search_count([
                ('id', '!=', rec.id),
                ('payroll_id', '=', rec.payroll_id.id),
                ('employee_id', '=', rec.employee_id.id),
            ])
            if dup:
                raise ValidationError('El empleado ya está en esta planilla.')