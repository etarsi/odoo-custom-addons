# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
from odoo.exceptions import UserError
from openpyxl import Workbook, load_workbook
import base64, re, pytz
from io import BytesIO
from datetime import datetime, date, time, timedelta

MONTHS_ES = {
    'ENERO': 1,
    'FEBRERO': 2,
    'MARZO': 3,
    'ABRIL': 4,
    'MAYO': 5,
    'JUNIO': 6,
    'JULIO': 7,
    'AGOSTO': 8,
    'SETIEMBRE': 9,
    'SEPTIEMBRE': 9,
    'OCTUBRE': 10,
    'NOVIEMBRE': 11,
    'DICIEMBRE': 12,
}


class HrAttendanceImportWizard(models.TransientModel):
    _name = 'hr.attendance.import.wizard'
    _description = 'Wizard: Importar Asistencias Eventuales'

    file = fields.Binary(string='Archivo Excel', required=True)
    filename = fields.Char(string='Nombre del archivo')

    # =========================
    #  Helpers de fechas / TZ
    # =========================

    def _get_tz(self):
        """Timezone a usar para convertir 07:00/17:00 a UTC."""
        tz_name = self.env.user.tz or 'America/Argentina/Buenos_Aires'
        return pytz.timezone(tz_name)

    def _make_dt_utc(self, day_date, hour, minute):
        """
        Devuelve un datetime naive en UTC correspondiente a day_date + hora local.
        Ej: 2025-11-10 07:00 (AR) -> 2025-11-10 10:00 (UTC) sin tzinfo.
        """
        tz = self._get_tz()
        local_dt = tz.localize(datetime.combine(day_date, time(hour, minute)))
        utc_dt = local_dt.astimezone(pytz.UTC)
        return utc_dt.replace(tzinfo=None)

    def _day_bounds_utc(self, day_date):
        """Rango [inicio, fin] del día en UTC, para buscar asistencias de ese día."""
        tz = self._get_tz()
        start_local = datetime.combine(day_date, time.min)
        end_local = datetime.combine(day_date, time.max)
        start_utc = tz.localize(start_local).astimezone(pytz.UTC).replace(tzinfo=None)
        end_utc = tz.localize(end_local).astimezone(pytz.UTC).replace(tzinfo=None)
        return start_utc, end_utc

    # =========================
    #  Helpers de Excel / Mes
    # =========================

    def _get_month_year_from_sheet(self, ws):
        """
        Intenta deducir mes y año del título de la pestaña.
        Ej: "11 NOVIEMBRE 2025" -> (11, 2025)
        """
        title = (ws.title or '').upper()

        # Año = primer 20xx que encuentre
        year_match = re.search(r'(20\d{2})', title)
        if not year_match:
            raise UserError(_('No se pudo deducir el año desde la pestaña "%s".') % ws.title)
        year = int(year_match.group(1))

        # Mes por nombre en español
        month = None
        for name, num in MONTHS_ES.items():
            if name in title:
                month = num
                break

        # Fallback: número en el título
        if not month:
            m = re.search(r'\b(1[0-2]|0?[1-9])\b', title)
            if m:
                month = int(m.group(1))

        if not month:
            raise UserError(_('No se pudo deducir el mes desde la pestaña "%s".') % ws.title)

        return month, year

    # =========================
    #  Helper: buscar empleado
    # =========================

    def _find_employee_by_cuil(self, cuil_raw):
        """
        Busca hr.employee comparando el CUIL del Excel con el campo dni del empleado.
        Se limpian puntos/guiones/espacios y se comparan solo dígitos.
        """
        if not cuil_raw:
            return None

        digits = re.sub(r'\D', '', str(cuil_raw))
        if not digits:
            return None

        # Se asume que dni en empleados está almacenado como solo dígitos
        emp = self.env['hr.employee'].search([('dni', '=', digits)], limit=1)
        if emp:
            return emp

        # Fallback por si en tu bd hay dni con formato
        emp = self.env['hr.employee'].search([('dni', '=', str(cuil_raw).strip())], limit=1)
        return emp

    # =========================
    #  Helper: P (Presente)
    # =========================

    def _upsert_attendance_presence(self, employee, day_date):
        if employee.type_shift == 'day':
            Attendance = self.env['hr.attendance']
            day_start_utc, day_end_utc = self._day_bounds_utc(day_date)

            attendances = Attendance.search([
                ('employee_id', '=', employee.id),
                ('check_in', '>=', day_start_utc),
                ('check_in', '<=', day_end_utc),
            ], order='check_in asc')

            # 1) Si hay alguna asistencia YA CERRADA (tiene check_out) -> no tocamos nada
            closed_att = attendances.filtered(lambda a: a.check_out)
            if closed_att:
                return

            # 2) Si hay abierta (tiene check_in pero no check_out) -> completar a las 17:00
            open_att = attendances.filtered(lambda a: not a.check_out)[:1]
            if open_att:
                checkout_dt = self._make_dt_utc(day_date, 17, 0)

                # Por si la marca fue después de las 17:00, evitar incoherencias
                if checkout_dt < open_att.check_in:
                    checkout_dt = open_att.check_in

                open_att.write({
                    'check_out': checkout_dt,
                    'type_income': 'P',
                })
                return

            # 3) No hay nada para ese día -> creamos de 07:00 a 17:00
            checkin_dt = self._make_dt_utc(day_date, 7, 0)
            checkout_dt = self._make_dt_utc(day_date, 17, 0)
        else:
            # Turno noche u otro: ajustar horarios según convenga
            Attendance = self.env['hr.attendance']
            day_start_utc, day_end_utc = self._day_bounds_utc(day_date)

            attendances = Attendance.search([
                ('employee_id', '=', employee.id),
                ('check_in', '>=', day_start_utc),
                ('check_in', '<=', day_end_utc),
            ], order='check_in asc')

            closed_att = attendances.filtered(lambda a: a.check_out)
            if closed_att:
                return

            open_att = attendances.filtered(lambda a: not a.check_out)[:1]
            if open_att:
                checkout_dt = self._make_dt_utc(day_date, 6, 0) + timedelta(days=1)
                if checkout_dt < open_att.check_in:
                    checkout_dt = open_att.check_in
                open_att.write({
                    'check_out': checkout_dt,
                    'type_income': 'P',
                })
                return

            checkin_dt = self._make_dt_utc(day_date, 20, 0)
            checkout_dt = self._make_dt_utc(day_date, 6, 0) + timedelta(days=1)

        Attendance.create({
            'employee_id': employee.id,
            'check_in': checkin_dt,
            'check_out': checkout_dt,
            'type_income': 'P',
            # Si usás estos campos:
            # 'employee_type': 'eventual',
            # 'employee_type_shift': 'day',
        })

    def _upsert_attendance_late(self, employee, day_date):
        if employee.type_shift == 'day':
            Attendance = self.env['hr.attendance']
            day_start_utc, day_end_utc = self._day_bounds_utc(day_date)

            attendances = Attendance.search([
                ('employee_id', '=', employee.id),
                ('check_in', '>=', day_start_utc),
                ('check_in', '<=', day_end_utc),
            ], order='check_in asc')

            closed_att = attendances.filtered(lambda a: a.check_out)
            if closed_att:
                return

            open_att = attendances.filtered(lambda a: not a.check_out)[:1]
            if open_att:
                checkout_dt = self._make_dt_utc(day_date, 17, 0)
                if checkout_dt < open_att.check_in:
                    checkout_dt = open_att.check_in
                open_att.write({
                    'check_out': checkout_dt,
                    'type_income': 'PT',
                })
                return

            checkin_dt = self._make_dt_utc(day_date, 7, 0)
            checkout_dt = self._make_dt_utc(day_date, 17, 0)
        else:
            Attendance = self.env['hr.attendance']
            day_start_utc, day_end_utc = self._day_bounds_utc(day_date)

            attendances = Attendance.search([
                ('employee_id', '=', employee.id),
                ('check_in', '>=', day_start_utc),
                ('check_in', '<=', day_end_utc),
            ], order='check_in asc')

            closed_att = attendances.filtered(lambda a: a.check_out)
            if closed_att:
                return

            open_att = attendances.filtered(lambda a: not a.check_out)[:1]
            if open_att:
                checkout_dt = self._make_dt_utc(day_date, 6, 0) + timedelta(days=1)
                if checkout_dt < open_att.check_in:
                    checkout_dt = open_att.check_in
                open_att.write({
                    'check_out': checkout_dt,
                    'type_income': 'PT',
                })
                return

            checkin_dt = self._make_dt_utc(day_date, 20, 0)
            checkout_dt = self._make_dt_utc(day_date, 6, 0) + timedelta(days=1)

        Attendance.create({
            'employee_id': employee.id,
            'check_in': checkin_dt,
            'check_out': checkout_dt,
            'type_income': 'PT',
        })

    def _upsert_attendance_absence(self, employee, day_date):
        """
        Marca de FALTA (F) para EVENTUALES.

        Crea una asistencia con check_in y check_out iguales al inicio del día.
        """
        if employee.type_shift == 'day':
            Attendance = self.env['hr.attendance']
            day_start_utc, day_end_utc = self._day_bounds_utc(day_date)

            attendances = Attendance.search([
                ('employee_id', '=', employee.id),
                ('check_in', '>=', day_start_utc),
                ('check_in', '<=', day_end_utc),
            ], order='check_in asc')

            if attendances:
                return

            checkin_dt = self._make_dt_utc(day_date, 0, 0)
        else:
            Attendance = self.env['hr.attendance']
            day_start_utc, day_end_utc = self._day_bounds_utc(day_date)

            attendances = Attendance.search([
                ('employee_id', '=', employee.id),
                ('check_in', '>=', day_start_utc),
                ('check_in', '<=', day_end_utc),
            ], order='check_in asc')

            if attendances:
                return

            checkin_dt = self._make_dt_utc(day_date, 20, 0)
        Attendance.create({
            'employee_id': employee.id,
            'check_in': checkin_dt,
            'check_out': checkin_dt,
            'type_income': 'F',
            'justification': 'Falta marcada desde importación',
        })

    # =========================
    #  Acción principal
    # =========================

    def action_import_attendance(self):
        """
        Importa asistencias desde el Excel de presentismo eventuales,
        para la pestaña actual (ej: "11 NOVIEMBRE 2025").

        - Column C = CUIL
        - Row 2 = números de día (1..30/31) desde la columna F.
        - Row 3+ = empleados.
        - Solo procesa código "P" (Presente) por ahora.
        """
        self.ensure_one()
        if not self.file:
            raise UserError(_('Debe adjuntar un archivo.'))

        try:
            data = base64.b64decode(self.file)
            wb = load_workbook(BytesIO(data), data_only=True)
        except Exception as e:
            raise UserError(_('No se pudo leer el archivo Excel: %s') % e)

        # Por simplicidad: usamos la hoja activa (donde está noviembre)
        ws = wb.active

        month, year = self._get_month_year_from_sheet(ws)

        max_row = ws.max_row
        max_col = ws.max_column

        # Configuración según tu archivo:
        DAY_HEADER_ROW = 2      # fila donde están los números de día
        EMP_START_ROW = 3       # primera fila de empleado
        CUIL_COL = 3            # columna C
        FIRST_DAY_COL = 5       # columna E

        # Mapear columna -> fecha (2025-11-01, etc.)
        col_to_date = {}
        for col in range(FIRST_DAY_COL, max_col + 1):
            val = ws.cell(row=DAY_HEADER_ROW, column=col).value
            if not val:
                continue
            try:
                day = int(val)
            except Exception:
                continue
            # si la celda no es un día válido, se salta
            if day <= 0 or day > 31:
                continue
            col_to_date[col] = date(year, month, day)

        if not col_to_date:
            raise UserError(_('No se encontraron columnas de días en la fila %s.') % DAY_HEADER_ROW)
        
        not_employees = []
        for row in range(EMP_START_ROW, max_row + 1):
            cuil_val = ws.cell(row=row, column=CUIL_COL).value

            # Si la fila está completamente vacía, la saltamos
            row_vals = [
                ws.cell(row=row, column=c).value
                for c in range(1, max_col + 1)
            ]
            if not any(row_vals):
                continue

            if not cuil_val:
                # Tiene algo pero sin CUIL -> lo ignoramos o levantamos error, a gusto
                # raise UserError(_('Fila %s tiene datos pero no CUIL.') % row)
                continue

            employee = self._find_employee_by_cuil(cuil_val)
            if not employee:
                not_employees.append(str(cuil_val))
                continue

            for col, day_date in col_to_date.items():
                raw = ws.cell(row=row, column=col).value
                if not raw:
                    continue

                code = str(raw).strip().upper()
                if code == 'P':
                    self._upsert_attendance_presence(employee, day_date)
                elif code == 'PT':
                    self._upsert_attendance_late(employee, day_date)
                elif code == 'F':
                    self._upsert_attendance_absence(employee, day_date)
                    
        # Que descargue un excel con los CUIL no encontrados, si querés
        if not_employees:
            archivo = "CUIL_no_encontrados.xlsx"
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.title = "CUIL no encontrados"
            ws_out.append(['CUIL no encontrado'])
            for cuil in not_employees:
                ws_out.append([cuil])

            output = BytesIO()
            wb_out.save(output)
            archivo_b64 = base64.b64encode(output.getvalue())

            # Crear un attachment
            attachment = self.env['ir.attachment'].create({
                'name': archivo,
                'type': 'binary',
                'datas': archivo_b64,
                'res_model': self._name,
                'res_id': self.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })

            # Devolver descarga directa y aviso
            return {
                'type': 'ir.actions.act_url',
                'url': '/web/content/%s?download=1' % attachment.id,
                'target': 'self',
            }

        # Si TODOS los empleados fueron encontrados
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Importación completada'),
                'message': _('Todos los empleados fueron encontrados.'),
                'type': 'success',
                'sticky': False,
                'timeout': 10000,
            }
        }
