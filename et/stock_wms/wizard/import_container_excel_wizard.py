# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
from odoo.exceptions import UserError
import base64
from io import BytesIO
from openpyxl import load_workbook


class ImportContainerExcelWizard(models.TransientModel):
    _name = 'import.container.excel.wizard'
    _description = 'Wizard: Importar Contenedores desde Excel'

    file = fields.Binary(string='Archivo Excel', required=True)
    nro_despacho = fields.Char(string='Número de Despacho', required=True)
    fecha_llegada = fields.Date(string='ETA')

    # ----------------- HELPERS -----------------

    def _normalize_code(self, value):
        """Convierte números/strings a código sin decimales tipo '56078'."""
        if value in (None, ''):
            return False
        # Si viene como float 56078.0 → '56078'
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(int(value))
        return str(value).strip()

    def _to_float(self, value):
        if value in (None, ''):
            return 0.0
        try:
            return float(value)
        except Exception:
            return 0.0

    def _extract_container_code(self, raw_text):
        """
        De algo como 'CONTAINER A#:JXLU4330628/A4250077459'
        devuelve 'JXLU4330628'.
        """
        if not raw_text:
            return False
        text = str(raw_text).strip()
        # Tomar lo que está después de ':'
        if ':' in text:
            text = text.split(':', 1)[1]
        # Tomar lo que está antes de '/'
        if '/' in text:
            text = text.split('/', 1)[0]
        text = text.strip()
        return text or False

    # ----------------- IMPORT PRINCIPAL -----------------
    def import_data(self):
        self.ensure_one()

        if not self.file:
            raise UserError(_("Debe adjuntar un archivo de Excel."))

        # Contenedor desde el que se abre el wizard
        container = self.env['container'].browse(self.env.context.get('active_id'))
        if not container:
            raise UserError(_("Este asistente debe abrirse desde un contenedor."))

        # Actualizar datos generales
        container.dispatch_number = self.nro_despacho
        if self.fecha_llegada:
            container.eta = self.fecha_llegada

        # Decodificar binario
        try:
            data = base64.b64decode(self.file)
        except Exception as e:
            raise UserError(_("No se pudo decodificar el archivo.\nDetalle técnico: %s") % e)

        # Cargar workbook .xlsx
        try:
            wb = load_workbook(BytesIO(data), data_only=True)
        except Exception as e:
            raise UserError(_("No se pudo leer el archivo Excel (.xlsx).\n"
                              "Asegúrese de que el archivo sea un .xlsx válido.\n"
                              "Detalle técnico: %s") % e)

        # Tomamos la hoja 'PACKING LIST' si existe, si no la activa
        sheet_name = 'PACKING LIST' if 'PACKING LIST' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        max_row = ws.max_row
        max_col = ws.max_column

        # -------- 1) Buscar fila de encabezados: ITEM CODE / CTNS --------
        header_row_idx = None
        for r in range(1, max_row + 1):
            row_vals = [
                (str(ws.cell(row=r, column=c).value).strip().upper()
                 if ws.cell(row=r, column=c).value is not None else '')
                for c in range(1, max_col + 1)
            ]
            if 'ITEM CODE' in row_vals and 'CTNS' in row_vals:
                header_row_idx = r
                break

        if not header_row_idx:
            raise UserError(_("No se encontró la fila de encabezados "
                              "('ITEM CODE' / 'CTNS') en el Excel."))

        # Mapear nombre de columna → índice
        header_row = [
            ws.cell(row=header_row_idx, column=c).value
            for c in range(1, max_col + 1)
        ]
        col_by_name = {}
        for idx, val in enumerate(header_row, start=1):
            if val:
                key = str(val).strip().upper()
                col_by_name[key] = idx

        def col(name):
            return col_by_name.get(name.upper())

        col_sb_code = col('SB CODE')
        col_barcode = col('BARCODE')
        col_ctns = col('CTNS')
        col_pcs = col('PCS')

        if not col_sb_code or not col_ctns or not col_pcs:
            raise UserError(_("No se encontraron las columnas 'SB CODE', 'CTNS' y/o 'PCS' en el encabezado."))

        # -------- 2) Buscar fila CONTAINER y TOTAL para cerrar el bloque --------
        container_row_idx = None
        total_row_idx = None

        for r in range(header_row_idx + 1, max_row + 1):
            row_vals_raw = [
                ws.cell(row=r, column=c).value
                for c in range(1, max_col + 1)
            ]
            row_text = [
                str(v).strip().upper()
                for v in row_vals_raw
                if v not in (None, '')
            ]
            if not row_text:
                # fila completamente vacía -> la ignoramos (no corta)
                continue

            if any('CONTAINER' in v for v in row_text):
                if not container_row_idx:
                    container_row_idx = r

            if any(v.startswith('TOTAL') for v in row_text):
                if not total_row_idx:
                    total_row_idx = r

            # Si ya tenemos ambas, cortamos
            if container_row_idx and total_row_idx:
                break

        # Determinar última fila de datos de detalle
        cut_points = [idx for idx in [container_row_idx, total_row_idx] if idx]
        if cut_points:
            last_data_row = min(cut_points) - 1
        else:
            # Fallback: hasta la última fila no vacía después del header
            last_data_row = header_row_idx
            for r in range(header_row_idx + 1, max_row + 1):
                row_vals = [
                    ws.cell(row=r, column=c).value
                    for c in range(1, max_col + 1)
                ]
                if any(v not in (None, '') for v in row_vals):
                    last_data_row = r
                else:
                    # primera fila totalmente vacía → cortamos
                    break

        # -------- 3) Extraer código de contenedor desde la fila CONTAINER --------
        if container_row_idx:
            row_vals_raw = [
                ws.cell(row=container_row_idx, column=c).value
                for c in range(1, max_col + 1)
            ]
            cont_cell = next(
                (v for v in row_vals_raw if v and 'CONTAINER' in str(v).upper()),
                None
            )
            cont_code = self._extract_container_code(cont_cell) if cont_cell else False
            if cont_code:
                # si querés, solo lo pisás si el name está vacío:
                # if not container.name:
                #     container.name = cont_code
                container.name = cont_code  # ← usa solo 'JXLU4330628' / 'JXLU4328277'

        # -------- 4) Construir las líneas del contenedor --------
        line_vals_to_create = []
        missing_codes = []

        for r in range(header_row_idx + 1, last_data_row + 1):
            # SB CODE → default_code del producto
            sb_val = ws.cell(row=r, column=col_sb_code).value if col_sb_code else None
            sb_code = self._normalize_code(sb_val)

            # Si no hay SB CODE, lo saltamos
            if not sb_code:
                continue

            product = self.env['product.product'].search(
                [('default_code', '=', sb_code)], limit=1
            )
            if not product:
                missing_codes.append(sb_code)
                continue

            qty_bultos = ws.cell(row=r, column=col_ctns).value if col_ctns else 0
            qty_cantidad = ws.cell(row=r, column=col_pcs).value if col_pcs else 0
            
            barcode_val = ws.cell(row=r, column=col_barcode).value if col_barcode else None
            barcode = self._normalize_code(barcode_val)

            vals = {
                'container': container.id,
                'product_id': product.id,
                'quantity_send': self._to_float(qty_cantidad),
                'bultos': self._to_float(qty_bultos),
                'quantity_picked': 0,
            }
            line_vals_to_create.append(vals)

        if missing_codes:
            missing_txt = ', '.join(sorted(set(missing_codes)))
            raise UserError(_(
                "Los siguientes 'SB CODE' no se encontraron como productos en Odoo "
                "(default_code):\n%s"
            ) % missing_txt)

        # Borramos las líneas actuales y cargamos las nuevas
        container.lines.unlink()
        for vals in line_vals_to_create:
            self.env['container.line'].create(vals)

        return {'type': 'ir.actions.act_window_close'}
