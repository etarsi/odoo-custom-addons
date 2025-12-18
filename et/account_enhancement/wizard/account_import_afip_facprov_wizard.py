# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError
import base64
import re
from io import BytesIO
from datetime import datetime, date
from openpyxl import load_workbook


class AccountImportAfipFacprovWizard(models.TransientModel):
    _name = 'account.import.afip.facprov.wizard'
    _description = 'Wizard: Importar Comprobantes AFIP (Proveedor) a Account Move'

    file = fields.Binary('Archivo Excel', required=True)
    company_id = fields.Many2one('res.company', 'Compañía', required=True,
                              default=lambda self: self.env.company)
    

    # ---------------- HELPERS ----------------
    def _extract_tipo_num(self, value):
        """
        Devuelve el número inicial del campo Tipo.
        Ejemplos:
        '1 - Factura A'  -> 1
        '11 - Factura C' -> 11
        '3 - Nota de Crédito A' -> 3
        11.0 -> 11
        """
        if value in (None, '', False):
            return False

        # Si Excel lo trae como número
        if isinstance(value, (int, float)):
            try:
                return int(value)
            except Exception:
                return False

        text = str(value).strip()

        # Caso típico: "11 - Factura C"
        m = re.match(r'^\s*(\d+)', text)
        return int(m.group(1)) if m else False

    def _to_float(self, v):
        if v in (None, '', False):
            return 0.0
        try:
            # openpyxl suele traer floats/ints; si viene string con coma, normalizamos
            if isinstance(v, str):
                v = v.replace('.', '').replace(',', '.') if v.count(',') == 1 else v
            return float(v)
        except Exception:
            return 0.0

    def _to_date(self, v):
        if not v:
            return False
        if isinstance(v, date) and not isinstance(v, datetime):
            return v
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, str):
            v = v.strip()
            # AFIP suele venir dd/mm/yyyy
            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
                try:
                    return datetime.strptime(v, fmt).date()
                except Exception:
                    continue
        return False

    def _norm_cuit(self, v):
        if not v:
            return False
        digits = re.sub(r"\D", "", str(v))
        return digits or False

    def _find_header_row(self, ws, max_scan=50):
        """
        Busca una fila que contenga al menos: 'Fecha', 'Tipo', 'Punto de Venta', 'Imp. Total'
        """
        must = {'FECHA', 'TIPO', 'PUNTO DE VENTA', 'NETO GRAVADO TOTAL', 'IMP. TOTAL'}
        for r in range(1, min(ws.max_row, max_scan) + 1):
            row = []
            for c in range(1, ws.max_column + 1):
                val = ws.cell(r, c).value
                if val:
                    row.append(str(val).strip().upper())
            if must.issubset(set(row)):
                return r
        return False

    def _build_col_map(self, ws, header_row):
        col_by_name = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(header_row, c).value
            if not val:
                continue
            col_by_name[str(val).strip().upper()] = c
        return col_by_name

    def _col(self, col_map, name):
        return col_map.get(name.strip().upper())

    def _get_purchase_vat_tax(self, rate, company_id):
        tax = self.env['account.tax'].search([
            ('type_tax_use', '=', 'purchase'),
            ('company_id', '=', company_id.id),
            ('amount', '=', rate),
            ('active', '=', True)], limit=1)
        return tax

    # ---------------- IMPORT PRINCIPAL ----------------
    def import_data(self):
        self.ensure_one()

        if not self.file:
            raise UserError(_("Debe adjuntar un archivo Excel (.xlsx)."))

        try:
            data = base64.b64decode(self.file)
        except Exception as e:
            raise UserError(_("No se pudo decodificar el archivo.\nDetalle: %s") % e)

        try:
            wb = load_workbook(BytesIO(data), data_only=True)
        except Exception as e:
            raise UserError(_("No se pudo leer el Excel. Asegúrese que sea .xlsx válido.\nDetalle: %s") % e)

        ws = wb[wb.sheetnames[0]]
        header_row = self._find_header_row(ws)
        if not header_row:
            raise UserError(_("No se encontró la fila de encabezados. Se esperaba un formato AFIP 'Mis Comprobantes Recibidos'."))

        # Columnas necesarias
        col_map = self._build_col_map(ws, header_row)
        c_fecha = self._col(col_map, 'Fecha')
        c_tipo = self._col(col_map, 'Tipo')
        c_pv = self._col(col_map, 'Punto de Venta')
        c_nro = self._col(col_map, 'Número Desde')
        c_cae = self._col(col_map, 'Cód. Autorización')
        c_emisor_doc = self._col(col_map, 'Nro. Doc. Emisor')
        c_emisor_name = self._col(col_map, 'Denominación Emisor')
        c_moneda = self._col(col_map, 'Moneda')
        c_tc = self._col(col_map, 'Tipo Cambio')
        c_total = self._col(col_map, 'Imp. Total')
        c_company = self._col(col_map, 'Nro. Doc. Receptor')

        if not all([c_fecha, c_tipo, c_pv, c_nro, c_emisor_doc, c_total]):
            raise UserError(_("Faltan columnas clave (Fecha/Tipo/Punto de Venta/Número Desde/Nro. Doc. Emisor/Imp. Total)."))

        # IVA bases + IVA
        iva_rates = [
            (0.0,  'Neto Grav. IVA 0%',   None),
            (2.5,  'Neto Grav. IVA 2,5%', 'IVA 2,5%'),
            (5.0,  'Neto Grav. IVA 5%',   'IVA 5%'),
            (10.5, 'Neto Grav. IVA 10,5%','IVA 10,5%'),
            (21.0, 'Neto Grav. IVA 21%',  'IVA 21%'),
            (27.0, 'Neto Grav. IVA 27%',  'IVA 27%'),
        ]
        c_ng = self._col(col_map, 'Neto No Gravado')
        c_ex = self._col(col_map, 'Op. Exentas')

        Move = self.env['account.move'].with_company(self.company_id)
        created_move_ids = []
        missing_taxes = set()
        move_type = 'in_invoice'
        # Recorremos filas de datos
        for r in range(header_row + 1, ws.max_row + 1):
            fecha = self._to_date(ws.cell(r, c_fecha).value)
            if not fecha:
                continue
            p_venta = ws.cell(r, c_pv).value
            num_fac = ws.cell(r, c_nro).value
            tipo = ws.cell(r, c_tipo).value
            tipo_code = self._extract_tipo_num(tipo)
            company_nif = ws.cell(r, c_company).value 
            company_actual = self.env.company.partner_id.vat
            tipo_cambio = ws.cell(r, c_tc).value 
            tipo_cambio = self._to_float(tipo_cambio) if tipo_cambio else 1.0
            emisor_cuit = self._norm_cuit(ws.cell(r, c_emisor_doc).value)
            emisor_name = ws.cell(r, c_emisor_name).value if c_emisor_name else False
            partner = self.env['res.partner'].search([('vat', '=', self._norm_cuit(emisor_cuit))], limit=1)   
            company_id = self.env['res.company'].search([('partner_id.vat', '=', self._norm_cuit(company_nif))], limit=1)
            fac_proveedor = self.env['account.move'].search([('name', 'ilike', num_fac), ('partner_id.vat', '=', self._norm_cuit(emisor_cuit)), ('company_id', '=', company_id.id)], limit=1)         
            tipo_comprobante = self.env['l10n_latam.document.type'].search([('code', '=', tipo_code)], limit=1)
            journal_id = self.env['account.journal'].search([('name', 'in', ['FACTURAS PROVEEDORES LAVALLE', 'FACTURAS PROVEEDORES DEPOSITO']), ('company_id', '=', company_id.id)], limit=1)
            currency = False
            moneda_symbol = ws.cell(r, c_moneda).value
            numero_documento = f"{int(p_venta):04d}-{int(num_fac):08d}"
            fila_no_registrada = ""
            # VALIDACIONES
            if not journal_id:
                raise ValidationError(_("No se encontró el diario para Facturas de Proveedores."))
            if not tipo_comprobante:
                raise ValidationError(_("Fila %s: tipo de comprobante inválido o no soportado: '%s'.") % (r, str(tipo)))
            if company_nif and company_actual and self._norm_cuit(company_nif) != self._norm_cuit(company_actual):
                raise ValidationError("Esta intentando verificar facturas que no corresponden a la compañía actual.")
            if not partner:
                fila_no_registrada += f"\n- Fila {r}: No se encontró el proveedor con CUIT '{emisor_cuit}'."
                continue
            if fac_proveedor:
                continue

            # Moneda
            if tipo_cambio == 1:
                currency = self.company_id.currency_id
            else:
                currency = self.env['res.currency'].search([('symbol', '=', moneda_symbol)], limit=1)
                if not currency:
                    raise ValidationError(_("Fila %s: no se encontró la moneda con símbolo '%s'.") % (r, ws.cell(r, c_moneda).value or ''))
            
            #INFORMACION DE LA FACTURA
            move_type = 'in_invoice'
            if tipo_comprobante.internal_type == 'credit_note':
                move_type = 'in_refund'

            #LINEAS 
            # Como separar el tip                
            # Construcción de líneas
            lines = []
            # Bases por alícuota
            for rate, neto_col_name, iva_col_name in iva_rates:
                c_neto = self._col(col_map, neto_col_name)
                if not c_neto:
                    continue
                neto = self._to_float(ws.cell(r, c_neto).value)
                if neto <= 0:
                    continue

                line = {
                    'name': ' ',
                    'account_id': journal_id.default_account_id.id,
                    'quantity': 1.0,
                    'price_unit': neto,
                }

                if rate and iva_col_name:
                    tax = self._get_purchase_vat_tax(rate, company_id)
                    if not tax:
                        missing_taxes.add(rate)
                    else:
                        line['tax_ids'] = [(6, 0, [tax.id])]

                lines.append((0, 0, line))

            # Neto no gravado
            if c_ng:
                tax = self.env['account.tax'].search([
                    ('type_tax_use', '=', 'purchase'),
                    ('company_id', '=', company_id.id),
                    ('name', 'ilike', 'IVA No Gravado'),
                    ('active', '=', True)], limit=1)
                if not tax:
                    raise ValidationError(_("No se encontró el impuesto para 'IVA No Gravado'."))
                ng = self._to_float(ws.cell(r, c_ng).value)
                if ng > 0:
                    lines.append((0, 0, {
                        'name': ' ',
                        'account_id': journal_id.default_account_id.id,
                        'quantity': 1.0,
                        'price_unit': ng,
                        'tax_ids': [(6, 0, [tax.id])]
                    }))

            # Exentas
            if c_ex:
                tax = self.env['account.tax'].search([
                    ('type_tax_use', '=', 'purchase'),
                    ('company_id', '=', company_id.id),
                    ('name', 'ilike', 'IVA Exento'),
                    ('active', '=', True)], limit=1)
                if not tax:
                    raise ValidationError(_("No se encontró el impuesto para 'IVA Exento'."))
                ex = self._to_float(ws.cell(r, c_ex).value)
                if ex > 0:
                    lines.append((0, 0, {
                        'name': ' ',
                        'account_id': journal_id.default_account_id.id,
                        'quantity': 1.0,
                        'price_unit': ex,
                        'tax_ids': [(6, 0, [tax.id])],
                    }))
            
            # Factura o nota de credito de letras C
            if tipo_comprobante.l10n_ar_letter == 'C':
                import_total = self._to_float(ws.cell(r, c_total).value) if c_total else 0.0
                lines.append((0, 0, {
                    'name': ' ',
                    'account_id': journal_id.default_account_id.id,
                    'quantity': 1.0,
                    'price_unit': import_total,
                }))

            vals = {
                'move_type': move_type,
                'company_id': self.company_id.id,
                'journal_id': journal_id.id,
                'l10n_latam_document_type_id': tipo_comprobante.id,
                'partner_id': partner.id,
                'invoice_date': fecha,
                'date': fecha,
                'currency_id': currency.id,
                'invoice_line_ids': lines,
                'computed_currency_rate': tipo_cambio,
                'l10n_ar_currency_rate': tipo_cambio,
                'l10n_latam_document_number': numero_documento,
            }

            move = Move.create(vals)
            created_move_ids.append(move.id)

        if not created_move_ids:
            response = {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Éxito',
                    'message': 'No se importó ningúna Factura de Proveedor Pendiente a importar.',
                    'type': 'info',
                    'sticky': False,
                    'timeout': 12000,
                }
            }    
        else:
            if not fila_no_registrada:
                response = {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': 'Éxito',
                        'message': f'Se importaron {len(created_move_ids)} - Facturas de Proveedor correctamente.',
                        'type': 'success',
                        'sticky': False,
                        'timeout': 12000,
                        'next': {
                            'type': 'ir.actions.act_window',
                            'name': _('Comprobantes Factura Importados'),
                            'res_model': 'account.move',
                            'view_mode': 'tree,form',
                            'domain': [('id', 'in', created_move_ids)],
                            'target': 'current',
                        }
                    }
                }
            else:
                response = {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': 'Atención',
                        'message': f'Se importaron {len(created_move_ids)} - Facturas de Proveedor. Sin embargo, Estas filas no se pudieron importar:\n{fila_no_registrada}',
                        'type': 'warning',
                        'sticky': True,
                        'timeout': 12000,
                        'next': {
                            'type': 'ir.actions.act_window',
                            'name': _('Comprobantes Factura Importados'),
                            'res_model': 'account.move',
                            'view_mode': 'tree,form',
                            'domain': [('id', 'in', created_move_ids)],
                            'target': 'current',
                        }
                    }
                }
        return response