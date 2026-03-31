# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
from odoo.exceptions import UserError
import base64, io, time, logging
from collections import defaultdict
from openpyxl import load_workbook

_logger = logging.getLogger(__name__)


class ImportSaleOrderMasiveWizard(models.TransientModel):
    _name = 'import.sale.order.masive.wizard'
    _description = 'Wizard: Importar Pedido de Venta Masivo'

    file = fields.Binary(string='Archivo Excel', required=True)
    sheet_name = fields.Char(string='Hoja', default='Sheet1', readonly=True)


    def _normalize(self, value):
        if value is None:
            return ''
        return str(value).strip()

    def _float_or_zero(self, value):
        if value in (None, '', False):
            return 0.0
        try:
            return float(value)
        except Exception:
            return 0.0

    def _load_excel_sheet(self):
        self.ensure_one()

        if not self.file:
            raise UserError(_('Debes cargar un archivo Excel.'))

        try:
            content = base64.b64decode(self.file)
            _logger.warning('Archivo Excel decodificado, tamaño en bytes: %s', len(content))
            wb = load_workbook(io.BytesIO(content), data_only=True) #data_only para leer valores calculados en vez de fórmulas, read_only para optimizar lectura de archivos grandes
            _logger.warning('Archivo Excel cargado en memoria, hojas disponibles: %s', ', '.join(wb.sheetnames))
        except Exception as e:
            raise UserError(_('No se pudo leer el archivo Excel:\n%s') % e)

        if self.sheet_name not in wb.sheetnames:
            raise UserError(_('No existe la hoja "%s". Hojas disponibles: %s') % (
                self.sheet_name, ', '.join(wb.sheetnames)
            ))

        return wb[self.sheet_name]

    def _read_sheet1_rows(self):
        """
        Lee la hoja Sheet1 con este formato:
        A Cliente
        B Dirección de entrega
        C Condición de venta
        D Plazos de pago
        E Términos y condiciones
        F Notas Internas
        G Descuento Rubro
        H Descuento
        I Producto
        J Cantidad
        M Compania forzada (opcional)

        Las filas heredan contexto de filas anteriores si vienen vacías.
        """
        self.ensure_one()
        ws = self._load_excel_sheet()

        rows = []
        current = {
            'cliente': '',
            'direccion_entrega': '',
            'condicion_venta': '',
            'plazos_pago': '',
            'terminos_condiciones': '',
            'notas_internas': '',
            'company_default': '',
            'discount_map': {},
            'default_discount': 0.0,
        }

        # Empieza en fila 2 porque fila 1 es encabezado
        for row_idx in range(2, ws.max_row + 1):
            cliente = self._normalize(ws[f'A{row_idx}'].value)
            direccion_entrega = self._normalize(ws[f'B{row_idx}'].value)
            condicion_venta = self._normalize(ws[f'C{row_idx}'].value)
            plazos_pago = self._normalize(ws[f'D{row_idx}'].value)
            terminos_condiciones = self._normalize(ws[f'E{row_idx}'].value)
            notas_internas = self._normalize(ws[f'F{row_idx}'].value)
            descuento_rubro = self._normalize(ws[f'G{row_idx}'].value)
            descuento = ws[f'H{row_idx}'].value
            producto_codigo = self._normalize(ws[f'I{row_idx}'].value)
            cantidad = ws[f'J{row_idx}'].value
            company_default = self._normalize(ws[f'M{row_idx}'].value)

            # Saltar fila vacía total
            if not any([
                cliente, direccion_entrega, condicion_venta, plazos_pago,
                terminos_condiciones, notas_internas, descuento_rubro,
                producto_codigo, cantidad, company_default
            ]):
                continue
            
            # Cuando arranca un bloque nuevo de cliente, reseteo mapa de descuentos
            if cliente:
                current['discount_map'] = {}
                current['default_discount'] = 0.0
            # Herencia de contexto cabecera
            if cliente:
                current['cliente'] = cliente
            if direccion_entrega:
                current['direccion_entrega'] = direccion_entrega
            if condicion_venta:
                current['condicion_venta'] = condicion_venta
            if plazos_pago:
                current['plazos_pago'] = plazos_pago
            if terminos_condiciones:
                current['terminos_condiciones'] = terminos_condiciones
            if notas_internas:
                current['notas_internas'] = notas_internas
            if company_default:
                current['company_default'] = company_default

            # Mapa de descuentos por rubro
            if descuento not in (None, '', False):
                descuento_val = self._float_or_zero(descuento)
                if descuento_rubro:
                    current['discount_map'][descuento_rubro.strip().upper()] = descuento_val
                else:
                    current['default_discount'] = descuento_val
            # Solo se importa si hay producto y cantidad
            if producto_codigo and cantidad not in (None, '', False):
                rows.append({
                    'excel_row': row_idx,
                    'cliente': current['cliente'],
                    'direccion_entrega': current['direccion_entrega'],
                    'condicion_venta': current['condicion_venta'],
                    'plazos_pago': current['plazos_pago'],
                    'terminos_condiciones': current['terminos_condiciones'],
                    'notas_internas': current['notas_internas'],
                    'producto_codigo': producto_codigo,
                    'cantidad': self._float_or_zero(cantidad),
                    'company_default': current['company_default'],
                    'discount_map': dict(current['discount_map']),
                    'default_discount': current['default_discount'],
                })
        if not rows:
            return []
        return rows

    def _build_master_data(self, rows):
        partner_names = set()
        shipping_names = set()
        product_codes = set()
        company_names = set()
        condicion_m2ms = set()
        payment_terms = set()

        for row in rows:
            if row['cliente']:
                partner_names.add(row['cliente'])
            if row['direccion_entrega']:
                shipping_names.add(row['direccion_entrega'])
            if row['producto_codigo']:
                product_codes.add(str(row['producto_codigo']).strip())
            if row['company_default']:
                company_names.add(row['company_default'])
            if row['condicion_venta']:
                condicion_m2ms.add(row['condicion_venta'])
            if row['plazos_pago']:
                payment_terms.add(row['plazos_pago'])

        partners = self.env['res.partner'].search([('name', 'in', list(partner_names))])
        products = self.env['product.product'].search([('default_code', 'in', list(product_codes))])
        condicion_m2ms = self.env['condicion.venta'].search([('name', 'in', list(condicion_m2ms))]) if condicion_m2ms else self.env['condicion.venta']
        companies = self.env['res.company'].search([('name', 'in', list(company_names))]) if company_names else self.env['res.company']
        payment_terms = self.env['account.payment.term'].search([('name', 'in', list(payment_terms))]) if payment_terms else self.env['account.payment.term']
        
        partner_map = {p.name.strip().upper(): p for p in partners if p.name}
        product_map = {p.default_code.strip().upper(): p for p in products if p.default_code}
        company_map = {c.name.strip().upper(): c for c in companies if c.name}
        condicion_m2m_map = {c.name.strip().upper(): c for c in condicion_m2ms if c.name}
        payment_term_map = {p.name.strip().upper(): p for p in payment_terms if p.name}

        return {
            'partner_map': partner_map,
            'product_map': product_map,
            'company_map': company_map,
            'condicion_m2m_map': condicion_m2m_map,
            'payment_term_map': payment_term_map,
        }

    def _resolve_company(self, row, master_data):
        company_name = (row.get('company_default') or '').strip()
        if company_name and company_name in master_data['company_map']:
            return master_data['company_map'][company_name]

        return self.env.company

    def _get_rubro_from_product(self, product):
        categ = product.categ_id
        if categ.parent_id:
            rubro = categ.parent_id.name
        else:
            rubro = categ.name if categ else ''
        return rubro

    def _build_group_key(self, row, partner, shipping, company, product):
        tipo = (row.get('condicion_venta') or '').strip().upper()
        rubro_real = self._get_rubro_from_product(product)

        if tipo == 'TIPO 3':
            return (
                partner.id,
                shipping.id if shipping else False,
                company.id,
                tipo,
            )

        return (
            partner.id,
            shipping.id if shipping else False,
            company.id,
            tipo,
            rubro_real,
        )
        
    def _get_discount_for_row(self, row, rubro_real):
        rubro_key = rubro_real.strip().upper() if rubro_real else ''
        discount_map = row.get('discount_map') or {}
        # 1) si el rubro tiene descuento específico, manda ese
        if rubro_key in discount_map:
            return self._float_or_zero(discount_map[rubro_key])
        # 2) si no tiene específico, usar descuento general del bloque
        return self._float_or_zero(row.get('default_discount', 0.0))
    
    def notifi_action_warning(self, message):
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Advertencia'),
                'type': 'warning',
                'message': message,
                'sticky': True,
            }
        }
    
    def _split_integer_qty(self, qty, parts):
        qty = float(qty or 0.0)
        if parts <= 0:
            return []

        if not qty.is_integer():
            raise UserError(_('La cantidad %s no es entera y no se puede dividir sin decimales.') % qty)

        qty_int = int(qty)
        base = qty_int // parts
        resto = qty_int % parts

        result = []
        for i in range(parts):
            result.append(base + (1 if i < resto else 0))
        return result
    
    def _resolve_shipping_partner(self, row, partner, shipping_cache=None):
        shipping_name = (row.get('direccion_entrega') or '').strip()
        if not shipping_name:
            return partner

        shipping_name_clean = shipping_name
        if ',' in shipping_name:
            parts = [part.strip() for part in shipping_name.split(',')]
            if len(parts) >= 2:
                shipping_name_clean = parts[-1]

        cache_key = (partner.id, shipping_name_clean)
        if shipping_cache is not None and cache_key in shipping_cache:
            return shipping_cache[cache_key]

        shipping = self.env['res.partner'].search([
            ('name', '=', shipping_name_clean),
            ('parent_id', '=', partner.id),
        ], limit=1)
        if not shipping:
            shipping = self.env['res.partner'].search([
                ('name', '=', shipping_name_clean),
            ], limit=1)
        if not shipping:
            shipping = self.env['res.partner'].search([
                ('name', 'ilike', shipping_name_clean),
                ('parent_id', '=', partner.id),
            ], limit=1)
        if not shipping:
            shipping = self.env['res.partner'].search([
                ('name', 'ilike', shipping_name_clean),
            ], limit=1)

        result = shipping or partner
        if shipping_cache is not None:
            shipping_cache[cache_key] = result
        return result
            
        

    def _prepare_order_vals(self, row, partner, shipping, company, global_discount, condicion_m2m, payment_term):
        note_parts = []
        if row.get('plazos_pago'):
            note_parts.append('Plazos de pago: %s' % row['plazos_pago'])
        if row.get('terminos_condiciones'):
            note_parts.append('Términos y condiciones: %s' % row['terminos_condiciones'])
        if row.get('notas_internas'):
            note_parts.append('Notas internas: %s' % row['notas_internas'])

        #SI ES PRODUCCION B
        if company.id == 1:
            condicion_m2m = self.env['condicion.venta'].search([('condicion', '=', 'C')], limit=1)

        return {
            'partner_id': partner.id,
            'partner_shipping_id': shipping.id if shipping else partner.id,
            'company_id': company.id,
            'note': '\n'.join(note_parts) if note_parts else False,
            'condicion_m2m': condicion_m2m.id if condicion_m2m else False,
            'global_discount': global_discount if global_discount > 0 else 0.0,
            'client_order_ref': False,
            'origin': 'IMPORT MASIVO',
            'payment_term_id': payment_term.id if payment_term else False,
        }

    def _prepare_order_line_vals(self, row, product):
        qty = row.get('cantidad') or 0.0
        if qty <= 0:
            return False, _('Fila %s: la cantidad debe ser mayor a cero.') % row['excel_row']

        vals = {
            'product_id': product.id,
            'name': product.get_product_multiline_description_sale() or product.display_name,
            'product_uom_qty': qty,
        }
        
        return vals, False 

    def action_import_file(self):
        self.ensure_one()
        rows = self._read_sheet1_rows()
        if not rows:
            return self.notifi_action_warning(_('No se encontraron líneas importables en la hoja %s.') % self.sheet_name)
        master_data = self._build_master_data(rows)
        sale_order = self.env['sale.order'].with_context(
            tracking_disable=True,
            mail_notrack=True,
            mail_create_nolog=True,
            mail_create_nosubscribe=True,
        )
        grouped = defaultdict(lambda: {
            'header': False,
            'lines': [],
            'discounts': set(),
            'tipo': '',
        })
        errors = []
        created_orders = self.env['sale.order']

        for row in rows:
            try:
                partner = master_data['partner_map'].get((row['cliente'] or '').strip().upper())
                if not partner:
                    errors.append(_('Fila %s: cliente no encontrado "%s".') % (
                        row['excel_row'], row['cliente']
                    ))
                    continue

                condicion_m2m = master_data['condicion_m2m_map'].get((row['condicion_venta'] or '').strip().upper())
                if row['condicion_venta'] and not condicion_m2m:
                    errors.append(_('Fila %s: condición de venta no encontrada "%s".') % (
                        row['excel_row'], row['condicion_venta']
                    ))
                    continue

                payment_term = master_data['payment_term_map'].get((row['plazos_pago'] or '').strip().upper())
                if row['plazos_pago'] and not payment_term:
                    errors.append(_('Fila %s: plazo de pago no encontrado "%s".') % (
                        row['excel_row'], row['plazos_pago']
                    ))
                    continue

                product = master_data['product_map'].get((row['producto_codigo'] or '').strip().upper())
                if not product:
                    errors.append(_('Fila %s: producto no encontrado "%s".') % (
                        row['excel_row'], row['producto_codigo']
                    ))
                    continue

                shipping = self._resolve_shipping_partner(row, partner)
                rubro_real = self._get_rubro_from_product(product)
                tipo = (row.get('condicion_venta') or '').strip().upper()
                descuento = self._get_discount_for_row(row, rubro_real)
                if tipo == 'TIPO 3':
                    target_companies =self._resolve_company(row, master_data)
                else:
                    # compañías destino de la línea
                    product_companies = product.company_ids
                    if product_companies:
                        target_companies = product_companies
                    else:
                        target_companies = self._resolve_company(row, master_data)

                if not target_companies:
                    errors.append(_('Fila %s: no se pudo determinar compañía para el producto "%s".') % (
                        row['excel_row'], product.display_name
                    ))
                    continue

                if len(target_companies) == 1:
                    qty_parts = [row.get('cantidad') or 0.0]
                else:
                    try:
                        qty_parts = self._split_integer_qty(row.get('cantidad') or 0.0, len(target_companies))
                    except Exception as e:
                        errors.append(_('Fila %s: error al dividir cantidad del producto "%s": %s') % (
                            row['excel_row'], product.display_name, str(e)
                        ))
                        continue

                for idx, company in enumerate(target_companies):
                    qty_part = qty_parts[idx]
                    if qty_part <= 0:
                        continue

                    row_part = dict(row)
                    row_part['cantidad'] = qty_part

                    line_vals, line_error = self._prepare_order_line_vals(row_part, product)
                    if line_error:
                        errors.append(line_error)
                        continue

                    group_key = self._build_group_key(row_part, partner, shipping, company, product)

                    if not grouped[group_key]['header']:
                        grouped[group_key]['header'] = self._prepare_order_vals(
                            row_part, partner, shipping, company, descuento, condicion_m2m, payment_term
                        )

                    grouped[group_key]['tipo'] = tipo
                    grouped[group_key]['lines'].append(line_vals)
                    grouped[group_key]['discounts'].add(descuento)
                    _logger.warning('Procesada fila %s para cliente "%s", producto "%s", compañía "%s", cantidad %s, descuento %s, grupo clave: %s',
                        row['excel_row'], partner.name, product.display_name, company.name, qty_part, descuento, group_key
                    )
            except Exception as e:
                errors.append(str(e))

        created_orders = self.env['sale.order']
        for group_key, data in sorted(grouped.items(), key=lambda x: (
            x[1]['header'].get('partner_id') or 0,
            x[1]['header'].get('company_id') or 0,
        )):
            _logger.warning('Procesando grupo de importación para partner_id=%s, company_id=%s, tipo=%s con %s líneas y descuentos: %s',
                data['header'].get('partner_id'), data['header'].get('company_id'), data['tipo'], len(data['lines']), ', '.join(str(d) for d in sorted(data['discounts']))
            )
            tipo = data['tipo']
            discounts = {d for d in data['discounts'] if d}

            if tipo == 'TIPO 3' and len(discounts) > 1:
                partner = self.env['res.partner'].browse(data['header']['partner_id'])
                errors.append(
                    _('No se puede importar el pedido para cliente "%s" porque tiene múltiples descuentos (%s) y condición de venta "Tipo 3".') % (
                        partner.name, ', '.join(str(d) for d in sorted(discounts))
                    )
                )
                continue

            global_discount = list(discounts)[0] if discounts else 0.0
            vals = dict(data['header'])
            vals['global_discount'] = global_discount if global_discount > 0 else 0.0
            vals['order_line'] = [(0, 0, line_vals) for line_vals in data['lines']]
            partner = self.env['res.partner'].browse(vals['partner_id'])
            attempts = 3
            
            while attempts > 0:
                try:
                    with self.env.cr.savepoint():
                        order = sale_order.create(vals)
                        created_orders |= order
                    self.env.cr.commit()
                    break
                
                except Exception as e:
                    msg = str(e).lower()
                    attempts -= 1
                    transient_error = ('could not serialize access due to concurrent update' in msg or 'deadlock detected' in msg)

                    if transient_error and attempts > 0:
                        _logger.warning('Reintentando pedido para cliente "%s" por error transitorio: %s', partner.name, str(e))
                        self.env.cr.rollback()
                        time.sleep(0.4)
                        continue

                    errors.append(_('Error al crear pedido para cliente "%s": %s') % (partner.name, str(e)))
                    self.env.cr.rollback()
                    break

        msg_ok = _('Importación finalizada. Pedidos creados: %s.') % len(created_orders)
        msg_error = '\n'.join(errors[:80]) if errors else ''
        
        if created_orders and errors:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Aviso'),
                    'message': _('Se han creado %s pedidos de venta.') % len(created_orders) + ('\n' + msg_error if msg_error else ''),
                    'type': 'warning',
                    'sticky': True,
                    'next': {
                        'type': 'ir.actions.act_window',
                        'name': _('Pedidos de Venta Importados'),
                        'res_model': 'sale.order',
                        'view_mode': 'tree,form',
                        'views': [(False, 'list'), (False, 'form')],
                        'domain': [('id', 'in', created_orders.ids)],
                        'target': 'current',
                    }
                }
            }

        elif created_orders and not errors:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Éxito'),
                    'message': msg_ok,
                    'sticky': True,
                    'type': 'success',
                    'next': {
                        'type': 'ir.actions.act_window',
                        'name': _('Pedidos de Venta Importados'),
                        'res_model': 'sale.order',
                        'view_mode': 'tree,form',
                        'views': [(False, 'list'), (False, 'form')],
                        'domain': [('id', 'in', created_orders.ids)],
                        'target': 'current',
                    }
                }
            }
        
        elif not created_orders and errors:
            return self.notifi_action_warning(msg_error)

    
    # for group_key, data in grouped.items():
    #     tipo = data['tipo']
    #     discounts = {d for d in data['discounts'] if d}

    #     # Si mezcla rubros y usa global_discount, no se puede representar más de un descuento
    #     if tipo == 'TIPO 3' and len(discounts) > 1:
    #         partner = self.env['res.partner'].browse(data['header']['partner_id'])
    #         errors.append(_('No se puede importar el pedido para cliente "%s" porque tiene múltiples descuentos (%s) y condición de venta "Tipo 3".') % (
    #             partner.name, ', '.join(str(d) for d in discounts)
    #         ))
    #         continue
    #     global_discount = list(discounts)[0] if discounts else 0.0
    #     vals= dict(data['header'])
    #     vals['global_discount'] = global_discount if global_discount > 0 else 0.0
    #     vals['order_line'] = [(0, 0, line_vals) for line_vals in data['lines']]
    #     try:
    #         with self.env.cr.savepoint():
    #             order = sale_order.create(vals)
    #             created_orders |= order
    #     except Exception as e:
    #         partner = self.env['res.partner'].browse(data['header']['partner_id'])  
    #         errors.append(_('Error al crear pedido para cliente "%s": %s') % (
    #             partner.name, str(e)
    #         ))
    #         continue